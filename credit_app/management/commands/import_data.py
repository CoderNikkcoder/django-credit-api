import os
from django.core.management.base import BaseCommand
from django.utils import timezone
from datetime import datetime
import openpyxl
from credit_app.models import Customer, Loan

class Command(BaseCommand):
    help = 'Import data from Excel files'
    
    def handle(self, *args, **options):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        
        customer_file = os.path.join(base_dir, 'customer_data.xlsx')
        loan_file = os.path.join(base_dir, 'loan_data.xlsx')
        
       
        self.stdout.write('Importing customers...')
        try:
            wb = openpyxl.load_workbook(customer_file)
            sheet = wb.active
            
            customer_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    customer_id, first_name, last_name, age, phone, monthly_salary, approved_limit = row
                    
                   
                    if not first_name or not last_name:
                        self.stdout.write(self.style.WARNING(f'Skipping customer {customer_id}: Missing name'))
                        continue
                    
                    
                    if isinstance(phone, float):
                        phone = int(phone)
                    
                    
                    customer, created = Customer.objects.get_or_create(
                        customer_id=customer_id,
                        defaults={
                            'first_name': first_name or 'Unknown',
                            'last_name': last_name or 'Customer',
                            'age': age or 0,
                            'phone_number': phone or 0,
                            'monthly_salary': monthly_salary or 0,
                            'approved_limit': approved_limit or 0
                        }
                    )
                    
                    if created:
                        customer_count += 1
                    else:
                        self.stdout.write(self.style.WARNING(f'Customer {customer_id} already exists'))
                        
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Error with customer {customer_id}: {str(e)}'))
                    continue
            
            self.stdout.write(self.style.SUCCESS(f'Successfully imported {customer_count} customers'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error importing customers: {str(e)}'))
        
        
        self.stdout.write('Importing loans...')
        try:
            wb = openpyxl.load_workbook(loan_file)
            sheet = wb.active
            
            loan_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    (customer_id, loan_id, loan_amount, tenure, interest_rate, 
                     monthly_payment, emis_paid, date_approved, end_date) = row
                    
                    
                    try:
                        customer = Customer.objects.get(customer_id=customer_id)
                    except Customer.DoesNotExist:
                        self.stdout.write(self.style.WARNING(f'Customer {customer_id} not found for loan {loan_id}'))
                        continue
                    
                    
                    if isinstance(date_approved, str):
                        date_approved = datetime.strptime(date_approved.split()[0], '%Y-%m-%d').date()
                    if isinstance(end_date, str):
                        end_date = datetime.strptime(end_date.split()[0], '%Y-%m-%d').date()
                    
                    
                    loan, created = Loan.objects.get_or_create(
                        loan_id=loan_id,
                        defaults={
                            'customer': customer,
                            'loan_amount': loan_amount or 0,
                            'tenure': tenure or 0,
                            'interest_rate': interest_rate or 0,
                            'monthly_payment': monthly_payment or 0,
                            'emis_paid_on_time': emis_paid or 0,
                            'date_of_approval': date_approved,
                            'end_date': end_date,
                            'start_date': date_approved
                        }
                    )
                    
                    if created:
                        loan_count += 1
                    else:
                        self.stdout.write(self.style.WARNING(f'Loan {loan_id} already exists'))
                        
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'Error importing loan {loan_id}: {str(e)}'))
                    continue
            
            self.stdout.write(self.style.SUCCESS(f'Successfully imported {loan_count} loans'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error importing loans: {str(e)}'))